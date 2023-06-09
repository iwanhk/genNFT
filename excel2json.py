import os
import json
import pandas as pd
from openpyxl import load_workbook
import copy

image_link = 'https://histore-isotop.oss-cn-beijing.aliyuncs.com/image/'
ignore = ["艺术品主图", "NFT信息录入", "版权信息->",
          "作品版权溯源->", "艺术品NFT藏品图片", "哈希值", "区块链信息查询->"]

template = {}
template["name"] = ""
template["description"] = ""
template["image"] = ""
template["attributes"] = []
template["descriptor"] = []

item = {
    "trait_type": "",
    "value": ""
}


def excel_to_json(file_path):
    # 加载Excel文件
    workbook = load_workbook(file_path, read_only=True)

    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 读取A和B列的内容
        LN = 1
        mode = 0

        json_data = copy.deepcopy(template)
        while True:
            key = sheet['A'+str(LN)].value
            value = sheet['B'+str(LN)].value
            if key == None:
                break
            LN += 1
            key = key.strip()
            if key[-1] == '：':
                key = key[:-1]
            key1 = key.split('、')
            if len(key1) > 1:
                key = key1[-1]

            if value == None or value == "/":
                value = ""
            if type(value) == type(0.1):
                value = str(int(value))
            # print(key, value)
            if key == '编号ID':
                json_data["image"] = image_link+value+'.jpg'
                continue
            if key == '作品名称':
                json_data["name"] = value
                continue
            if key == '描述':
                json_data["description"] = value
                continue
            if key == "作品版权溯源->":
                mode = 1
                continue
            if key in ignore:
                continue
            if value == "":
                continue

            if mode == 0:
                i = copy.deepcopy(item)
                i["trait_type"] = key
                i["value"] = value
                json_data["attributes"].append(i)
            else:
                i = copy.deepcopy(item)
                i["trait_type"] = key
                i["value"] = value
                json_data["descriptor"].append(i)

        # 将JSON数据保存到文件中
        with open(f'zhongyou/{sheet_name}.json', 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file,
                      ensure_ascii=False, indent=4)

        print(f"{sheet_name}.json 文件已生成")
        del json_data
        del i


if __name__ == "__main__":
    # 替换为您的Excel文件路径
    excel_file_path = "zhongyou/2023-6月-加密艺术品.xlsx"
    excel_to_json(excel_file_path)
